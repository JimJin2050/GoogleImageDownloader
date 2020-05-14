# -*- coding:utf-8 -*-
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image


class ExcelHandler(object):
    def __init__(self, filename):
        self.workbook = load_workbook(filename)
        self.sheet = self.workbook.active

    @property
    def rows(self):
        return self.sheet.max_row

    @property
    def columns(self):
        return self.sheet.max_column

    def create_sheet(self):
        return self.sheet.create_sheet(0)

    def get_cell_value(self, row, col):
        return self.sheet.cell(row=row, column=col).value

    def set_cell_value(self, row, col, val):
        self.sheet.cell(row=row, column=col).value = val

    def cell(self, row, col):
        return self.sheet.cell(row=row, column=col)

    def save(self, filename):
        self.workbook.save(filename)

    def add_image(self, filename, cell):
        img = Image(filename)
        self.sheet.add_image(img, cell)

    def add_txt(self, val_list):
        self.sheet.append(val_list)

    def close(self):
        self.workbook.close()


if __name__ == "__main__":
    pass
