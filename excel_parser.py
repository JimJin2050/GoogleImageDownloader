# -*- coding:utf-8 -*-
from openpyxl import load_workbook


class ExcelParser(object):
    def __init__(self, filename):
        self.filename = filename


    @property
    def workbook(self):
        return load_workbook(filename=self.filename)

    @property
    def sheet(self):
        return self.workbook.active

    @property
    def rows(self):
        return self.sheet.max_row

    @property
    def columns(self):
        return self.sheet.max_column

    def get_keywords(self):
        keywords = []
        for r in range(2, self.rows + 1):
            col1 = self.sheet.cell(row=r, column=1).value
            col2 = self.sheet.cell(row=r, column=2).value
            keywords.append("{} {}".format(col1, col2))
        return keywords

    def close(self):
        self.workbook.close()


if __name__ == "__main__":
    pass
