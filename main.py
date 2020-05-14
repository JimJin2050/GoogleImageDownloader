# -*- coding:utf-8 -*-
import os
import requests
import time
import yaml
import base64
from datetime import datetime

from openpyxl.drawing.image import Image
from excel_parser import ExcelParser
from excel_handler import ExcelHandler
from my_driver import MyDriver
from google_page import GooglePage

setting_file = "settings.yaml"


def get_img_folder(idx, keyword, platform="macos"):
    keyword = keyword.replace("/", "")
    cur_dir = os.path.dirname(os.path.abspath(__file__))
    downloads_dir = "{}/{}".format(cur_dir, "downloads")
    img_folder = "{}/{}".format(downloads_dir, "{}-{}".format(idx, keyword))
    if platform == "windows":
        img_folder = create_folder_on_windows(img_folder)
    else:
        if not os.path.exists(img_folder):
            os.system(r"mkdir '{}'".format(img_folder))
            time.sleep(1)
    return img_folder


def create_folder_on_windows(my_folder):
    my_folder = my_folder.replace("/", "\\")
    if not os.path.exists(my_folder):
        os.system(r'mkdir "{}"'.format(my_folder))
        time.sleep(1)
    return my_folder


def generate_folder_by_time(rela_root, platform="macos"):
    folder = "{}/{}".format(rela_root, datetime.now().strftime("%Y%m%d%H%M%S"))
    if platform == "windows":
        folder = create_folder_on_windows(folder)
    else:
        if not os.path.exists(folder):
            os.system(r"mkdir '{}'".format(folder))
            time.sleep(1)
    return folder


def get_image_by_url(dir, img_url, index, platform="macos"):
    img_file = "{}/{}.png".format(dir, index)
    if platform == "windows":
        img_file = img_file.replace("/", "\\")
    r = requests.get(img_url, stream=True, timeout=60, verify=True)
    if r.status_code == 200:
        with open(img_file, 'wb') as image_file:
            image_file.write(r.content)
            print("Done")
    r.close()


def download_img(dir, img_data, index, platform="macos"):
    img_file = "{}/{}.png".format(dir, index)
    if platform == "windows":
        img_file = img_file.replace("/", "\\")
    with open(img_file, 'wb') as png_file:
        try:
            if img_data.startswith("http"):
                get_image_by_url(dir, img_data, index, platform)
            else:
                data = img_data.split(",")[-1]
                decoded = base64.b64decode(data)
                png_file.write(decoded)
        except TypeError as e:
            print(e)
            pass
    return img_file


def init_driver(dr, imp_time=30, load_time=30, script_time=60):
    dr.implicitly_wait(imp_time)
    dr.set_page_load_timeout(load_time)
    dr.set_script_timeout(script_time)
    return dr


def get_settings():
    constants_path = os.path.join(os.path.dirname(__file__), setting_file)
    with open(constants_path, 'r', encoding="utf-8") as f:
        file_data = f.read()
    return yaml.safe_load(file_data)


def save_to_excel(excel_handler, img_file, row, keyword):
    #excel_handler = ExcelHandler()
    cell1 = str(excel_handler.sheet.cell(row, 1)).split(".")[1].strip(">")
    excel_handler.sheet.column_dimensions[cell1[0:1]].width = 36.0
    excel_handler.sheet.row_dimensions[int(cell1[1:])].height = 180.0
    excel_handler.set_cell_value(row, 1, keyword)
    cell2 = str(excel_handler.sheet.cell(row, 2)).split(".")[1].strip(">")
    print(row, cell2)
    excel_handler.sheet.column_dimensions[cell2[0:1]].width = 36.0
    excel_handler.sheet.row_dimensions[int(cell2[1:])].height = 180.0
    excel_handler.add_image(img_file, cell2)
    excel_handler.save(output_excel_file)


if __name__ == "__main__":
    #path = r"C:\Users\JIM\PycharmProjects\GoogleImageDownloader\data\results.xlsx"
    # data = get_settings()
    # platform = data["os"]
    # start_index = data.get("start_index")
    # excel_file = data.get("excel_file")
    # output_excel_file = data.get("output_excel_file")
    # num_to_download = data.get("num_to_download")
    # headless = data.get("headless")
    # excel = ExcelHandler()
    # excel.sheet.column_dimensions['B'].width = 36.0
    # excel.sheet.row_dimensions[1].height = 180.0
    # excel.set_cell_value(1, 1, "test")
    # cell = str(excel.cell(1, 2)).split(".")[1].strip(">")
    # excel.add_image("1.png", cell)
    # excel.save(output_excel_file)

    # Get settings
    data = get_settings()
    platform = data["os"]
    start_index = data.get("start_index")
    excel_file = data.get("excel_file")
    output_excel_file = data.get("output_excel_file")
    num_to_download = data.get("num_to_download")
    headless = data.get("headless")

    excel_handler = ExcelHandler(output_excel_file)

    # Get keywords
    excel_parser = ExcelParser(excel_file)
    keywords = excel_parser.get_keywords()

    # Init a webdriver instance
    driver = MyDriver.chrome_driver(headless)
    driver = init_driver(driver)

    # Go to image search page
    search_page = GooglePage(driver)
    search_page.go_to_img_search_page()
    row = 1
    try:
        for idx, keyword in enumerate(keywords):
            if idx >= start_index - 1:
                print("Number {}: {}".format(idx + 1, keyword))
                # generate folder to store images
                img_folder = get_img_folder(idx + 1, keyword, platform)
                # folder = generate_folder_by_time(img_folder, platform)
                folder = img_folder

                # Search images by keyword
                search_page.search_img_by_keyword(keyword)
                # Get images data from page
                img_data_list = search_page.get_images(keyword, num_to_download)

                # Downloads images and store into folder

                for ind, img_data in enumerate(img_data_list):
                    img_file = download_img(folder, img_data, ind + 1, platform)
                    save_to_excel(excel_handler, img_file, row, keyword)
                    row += 1



    except Exception as e:
        print(e)
    finally:
        # Quit the webdriver instance
        excel_handler.close()
        driver.quit()
        pass
